from openpyxl import Workbook
import random

# Criação do workbook e da planilha
workbook = Workbook()
sheet = workbook.active
sheet.title = 'estoque'

# Cabeçalhos da planilha
headers = ['Nome do Produto', 'Valor do Fornecedor', 'Lucratividade (%)', 'Quantidade']
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)

# Função para gerar nomes de produtos aleatórios
def gerar_nome_produto():
    prefixos = ['Super', 'Mega', 'Ultra', 'Power', 'Eco', 'Max']
    tipos = ['Widget', 'Gadget', 'Device', 'Tool', 'Instrument', 'Appliance']
    sufixos = ['Plus', 'Pro', 'X', '2000', 'Prime', 'Elite']
    return f'{random.choice(prefixos)} {random.choice(tipos)} {random.choice(sufixos)}'

# Número de produtos a serem gerados
num_produtos = 50

# Preenchimento da planilha com dados aleatórios
for row_num in range(2, num_produtos + 2):
    nome_produto = gerar_nome_produto()
    valor_fornecedor = round(random.uniform(10.0, 500.0), 2)
    lucratividade = random.randint(10, 100)
    quantidade = random.randint(1, 100)
    
    sheet.cell(row=row_num, column=1, value=nome_produto)
    sheet.cell(row=row_num, column=2, value=valor_fornecedor)
    sheet.cell(row=row_num, column=3, value=lucratividade)
    sheet.cell(row=row_num, column=4, value=quantidade)

# Salvar o arquivo
file_path = 'estoque.xlsx'
workbook.save(file_path)
print(f'Arquivo salvo em {file_path}')
